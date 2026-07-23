import csv
import io
import json
import os
import re
from collections import defaultdict
from datetime import datetime
from decimal import Decimal
from typing import Any, Dict, Iterable, List, Optional

from fastapi import HTTPException, status
from sqlalchemy.orm import Session

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None

try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None

try:
    from PyPDF2 import PdfReader
except ImportError:
    try:
        from pypdf import PdfReader
    except ImportError:
        PdfReader = None

try:
    from rapidocr_onnxruntime import RapidOCR
except ImportError:
    RapidOCR = None

from src.models.documento import Carpeta, Documento
from src.models.empresa import Empresa
from src.models.licitacion import Licitacion, RequisitoChecklist

OCR_ENGINE: Optional[Any] = None

DEFAULT_REQUIRED_DOCUMENTS = [
    {
        "key": "rup",
        "nombre": "RUP vigente",
        "descripcion": "Registro unico de proponentes actualizado y legible.",
        "keywords": ["rup", "registro unico de proponentes", "registro unico"],
        "categoria": "habilitante",
        "obligatorio": True,
    },
    {
        "key": "camara",
        "nombre": "Certificado de existencia y representacion legal",
        "descripcion": "Camara de comercio o equivalente, vigente.",
        "keywords": ["existencia", "representacion legal", "camara de comercio", "certificado mercantil"],
        "categoria": "juridico",
        "obligatorio": True,
    },
    {
        "key": "rut",
        "nombre": "RUT",
        "descripcion": "Registro unico tributario actualizado.",
        "keywords": ["rut", "registro unico tributario"],
        "categoria": "tributario",
        "obligatorio": True,
    },
    {
        "key": "financiero",
        "nombre": "Estados financieros",
        "descripcion": "Estados financieros recientes, firmados cuando aplique.",
        "keywords": ["estados financieros", "balance", "estado de resultados", "revisor fiscal"],
        "categoria": "financiero",
        "obligatorio": True,
    },
    {
        "key": "presentacion",
        "nombre": "Carta de presentacion",
        "descripcion": "Carta de presentacion o propuesta firmada.",
        "keywords": ["carta de presentacion", "carta propuesta", "presentacion de la oferta"],
        "categoria": "juridico",
        "obligatorio": True,
    },
    {
        "key": "cedula",
        "nombre": "Cedula representante legal",
        "descripcion": "Copia de la cedula del representante legal.",
        "keywords": ["cedula", "documento de identidad", "representante legal"],
        "categoria": "juridico",
        "obligatorio": True,
    },
    {
        "key": "experiencia",
        "nombre": "Experiencia habilitante",
        "descripcion": "Certificaciones o soportes de experiencia.",
        "keywords": ["experiencia", "certificacion", "contrato", "acta de liquidacion"],
        "categoria": "tecnico",
        "obligatorio": True,
    },
    {
        "key": "garantia",
        "nombre": "Garantia de seriedad",
        "descripcion": "Garantia o poliza de seriedad de la oferta.",
        "keywords": ["seriedad de la oferta", "garantia de seriedad", "poliza de seriedad"],
        "categoria": "financiero",
        "obligatorio": False,
    },
    {
        "key": "parafiscales",
        "nombre": "Seguridad social y parafiscales",
        "descripcion": "Soportes de aportes a seguridad social y parafiscales.",
        "keywords": ["seguridad social", "parafiscales", "salud", "pension"],
        "categoria": "juridico",
        "obligatorio": True,
    },
    {
        "key": "anexos",
        "nombre": "Anexos tecnicos",
        "descripcion": "Anexos tecnicos, fichas y formatos del pliego.",
        "keywords": ["anexo tecnico", "ficha tecnica", "especificaciones", "anexo"],
        "categoria": "tecnico",
        "obligatorio": False,
    },
]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip().lower()


# ============================================
# INDICADORES FINANCIEROS / CAPACIDAD ORGANIZACIONAL (RUP)
# ============================================

def _valor_pattern(label: str) -> str:
    """Construye un patron tolerante a como PyMuPDF/OCR aplanan las tablas del RUP: entre la
    etiqueta del indicador y su valor puede haber espacios, puntos de relleno, guiones o incluso
    un salto de linea (cuando la columna de valores queda separada de la de etiquetas)."""
    return rf"{label}[^\d\n]{{0,40}}\n?[^\d\n]{{0,40}}([\d]+(?:[.,]\d+)?)"


_LABEL_LIQUIDEZ = r"[IÍ]NDICE\s+DE\s+LIQUIDEZ"
_LABEL_ENDEUDAMIENTO = r"[IÍ]NDICE\s+DE\s+ENDEUDAMIENTO"
_LABEL_COBERTURA_INTERESES = r"RAZ[OÓ]N\s+DE\s+COBERTURA\s+DE\s+INTERESES"
_LABEL_RENT_PATRIMONIO = r"RENTABILIDAD\s+DEL\s+PATRIMONIO"
_LABEL_RENT_ACTIVO = r"RENTABILIDAD\s+DEL\s+ACTIVO"

INDICADORES_FINANCIEROS_DEF = {
    "indice_liquidez": {
        "nombre": "Índice de liquidez",
        "label": _LABEL_LIQUIDEZ,
        "patron": _valor_pattern(_LABEL_LIQUIDEZ),
        "operador": ">=",
        "unidad": "ratio",
    },
    "indice_endeudamiento": {
        "nombre": "Índice de endeudamiento",
        "label": _LABEL_ENDEUDAMIENTO,
        "patron": _valor_pattern(_LABEL_ENDEUDAMIENTO),
        "operador": "<=",
        "unidad": "porcentaje",
    },
    "razon_cobertura_intereses": {
        "nombre": "Razón de cobertura de intereses",
        "label": _LABEL_COBERTURA_INTERESES,
        "patron": _valor_pattern(_LABEL_COBERTURA_INTERESES),
        "operador": ">=",
        "unidad": "ratio",
    },
    "rentabilidad_patrimonio": {
        "nombre": "Rentabilidad del patrimonio",
        "label": _LABEL_RENT_PATRIMONIO,
        "patron": _valor_pattern(_LABEL_RENT_PATRIMONIO),
        "operador": ">=",
        "unidad": "porcentaje",
    },
    "rentabilidad_activo": {
        "nombre": "Rentabilidad del activo",
        "label": _LABEL_RENT_ACTIVO,
        "patron": _valor_pattern(_LABEL_RENT_ACTIVO),
        "operador": ">=",
        "unidad": "porcentaje",
    },
}

# Patrones para intentar detectar, en el pliego, el minimo/maximo exigido para cada indicador.
# Son heuristicos (el texto del pliego varia mucho): si no encuentran nada, el usuario lo
# digita a mano y ese valor manual siempre tiene prioridad.
_REQUISITO_KEYWORDS = (
    r"(?:mayor\s+o\s+igual\s+a|superior\s+o\s+igual\s+a|igual\s+o\s+superior\s+a|"
    r"m[ií]nimo|no\s+inferior\s+a|no\s+menor\s+a|inferior\s+o\s+igual\s+a|"
    r"menor\s+o\s+igual\s+a|igual\s+o\s+inferior\s+a|m[áa]ximo|no\s+superior\s+a|"
    r"hasta|>=|=>|≥|<=|=<|≤)"
)


def _parse_decimal_co(valor: Any) -> Optional[float]:
    """Convierte un numero a float aceptando formato colombiano ('0,21' / '1.234,56')
    y tambien el formato con punto decimal, por si se digita a mano en el formulario."""
    if valor is None:
        return None
    texto = str(valor).strip()
    if not texto:
        return None

    tiene_coma = "," in texto
    tiene_punto = "." in texto

    try:
        if tiene_coma and tiene_punto:
            return float(texto.replace(".", "").replace(",", "."))
        if tiene_coma:
            return float(texto.replace(",", "."))
        return float(texto)
    except ValueError:
        return None


def _normalizar_porcentaje(valor: float) -> float:
    """La Camara de Comercio no es consistente en como imprime los porcentajes en el RUP/pliego:
    algunas certificaciones ya traen la fraccion (p. ej. '0,17' = 17%) y otras traen el numero
    completo (p. ej. '32.45' = 32.45%). El resto del pipeline (entrada manual e
    IndicadoresFinancieros.jsx) siempre guarda/espera la fraccion (0-1), asi que solo se divide
    entre 100 cuando el valor leido es mayor a 1 (es decir, cuando claramente NO es ya una
    fraccion)."""
    return valor / 100 if valor > 1 else valor


def extraer_indicadores_financieros_rup(texto: str) -> Dict[str, Dict[str, Any]]:
    """Lee del RUP los indicadores financieros/de capacidad organizacional (formato estandar
    de Camara de Comercio: 'INDICADOR : valor')."""
    if not texto:
        return {}

    resultado = {}
    for key, definicion in INDICADORES_FINANCIEROS_DEF.items():
        match = re.search(definicion["patron"], texto, re.IGNORECASE)
        if not match:
            continue
        valor = _parse_decimal_co(match.group(1))
        if valor is None:
            continue
        if definicion["unidad"] == "porcentaje":
            valor = _normalizar_porcentaje(valor)
        resultado[key] = {
            "key": key,
            "nombre": definicion["nombre"],
            "valor": valor,
            "valor_texto": match.group(1).strip(),
        }
    return resultado


def extraer_requisitos_financieros_pliego(texto: str) -> Dict[str, Dict[str, Any]]:
    """Intento heuristico de detectar, en el pliego, el minimo/maximo exigido por indicador.
    Es un punto de partida (OCR); el usuario siempre puede corregirlo a mano.

    Primero busca la frase explicita de minimo/maximo ("mayor o igual a", ">=", etc.). Si el
    pliego no usa ninguna de esas frases (muchos solo traen una tabla plana "INDICADOR : valor",
    igual que el RUP), cae al mismo patron tolerante que se usa para leer el RUP."""
    if not texto:
        return {}

    resultado = {}
    for key, definicion in INDICADORES_FINANCIEROS_DEF.items():
        match = re.search(
            rf"{definicion['label']}[^\n]{{0,120}}?{_REQUISITO_KEYWORDS}[^\d]{{0,15}}([\d.,]+)",
            texto,
            re.IGNORECASE,
        )
        if not match:
            match = re.search(definicion["patron"], texto, re.IGNORECASE)
        if not match:
            continue

        valor = _parse_decimal_co(match.group(1))
        if valor is None:
            continue
        if definicion["unidad"] == "porcentaje":
            valor = _normalizar_porcentaje(valor)
        resultado[key] = {"key": key, "valor": valor, "valor_texto": match.group(1).strip()}
    return resultado


def comparar_indicadores_financieros(
    rup_texto: Optional[str],
    pliego_texto: Optional[str],
    requeridos_manual: Optional[Dict[str, Any]] = None,
    rup_manual: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    """Combina lo detectado en el RUP y el pliego con lo que el usuario haya digitado a mano
    (manual siempre tiene prioridad sobre el OCR) y marca si cumple (verde) o no (rojo).
    La tabla siempre trae los 5 indicadores estandar, tengan dato o no."""
    valores_rup_ocr = extraer_indicadores_financieros_rup(rup_texto)
    requeridos_ocr = extraer_requisitos_financieros_pliego(pliego_texto)
    requeridos_manual = requeridos_manual or {}
    rup_manual = rup_manual or {}

    filas = []
    for key, definicion in INDICADORES_FINANCIEROS_DEF.items():
        rup_manual_valor = _parse_decimal_co(rup_manual.get(key))
        rup_ocr_item = valores_rup_ocr.get(key)

        if rup_manual_valor is not None:
            valor_rup = rup_manual_valor
            fuente_rup = "manual"
        elif rup_ocr_item is not None:
            valor_rup = rup_ocr_item["valor"]
            fuente_rup = "rup_ocr"
        else:
            valor_rup = None
            fuente_rup = None

        requerido_manual_valor = _parse_decimal_co(requeridos_manual.get(key))
        ocr_item = requeridos_ocr.get(key)

        if requerido_manual_valor is not None:
            valor_requerido = requerido_manual_valor
            fuente_requerido = "manual"
        elif ocr_item is not None:
            valor_requerido = ocr_item["valor"]
            fuente_requerido = "pliego_ocr"
        else:
            valor_requerido = None
            fuente_requerido = None

        cumple = None
        if valor_rup is not None and valor_requerido is not None:
            if definicion["operador"] == ">=":
                cumple = valor_rup >= valor_requerido
            else:
                cumple = valor_rup <= valor_requerido

        filas.append(
            {
                "key": key,
                "nombre": definicion["nombre"],
                "operador": definicion["operador"],
                "unidad": definicion["unidad"],
                "valor_rup": valor_rup,
                "fuente_rup": fuente_rup,
                "valor_requerido": valor_requerido,
                "fuente_requerido": fuente_requerido,
                "cumple": cumple,
            }
        )

    return filas


def limpiar_codigo(codigo: str) -> str:
    if codigo is None:
        return ""
    return re.sub(r"\D", "", str(codigo))


def formatear_codigo(codigo: str) -> str:
    codigo = limpiar_codigo(codigo)
    if len(codigo) == 8:
        return f"{codigo[0:2]} {codigo[2:4]} {codigo[4:6]} {codigo[6:8]}"
    return codigo


def obtener_codigos_desde_texto(texto: str) -> List[str]:
    lineas = [line.strip() for line in re.split(r"[\n,;]+", texto or "") if line.strip()]
    codigos = []
    for item in lineas[:10]:
        limpio = limpiar_codigo(item)
        if limpio:
            codigos.append(limpio)
    return codigos


def _get_ocr_engine() -> Optional[Any]:
    global OCR_ENGINE
    if RapidOCR is None:
        return None
    if OCR_ENGINE is None:
        OCR_ENGINE = RapidOCR()
    return OCR_ENGINE


def _ocr_image_bytes(image_bytes: bytes) -> str:
    engine = _get_ocr_engine()
    if engine is None:
        return ""
    try:
        result, _ = engine(image_bytes)
    except Exception:
        return ""

    textos = []
    for block in result or []:
        if isinstance(block, (list, tuple)) and len(block) >= 2 and block[1]:
            textos.append(str(block[1]).strip())
    return "\n".join([line for line in textos if line])


def _extract_pdf_text_with_reader(file_bytes: bytes) -> Dict[str, Any]:
    if PdfReader is None:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=(
                "No se pudo procesar el PDF porque faltan librerias de lectura. "
                "Instala PyMuPDF o PyPDF2/pypdf en el backend."
            ),
        )

    reader = PdfReader(io.BytesIO(file_bytes))
    pages_text = []
    for page in reader.pages:
        pages_text.append((page.extract_text() or "").strip())

    combined = "\n".join([page_text for page_text in pages_text if page_text.strip()])
    return {
        "text": combined,
        "page_count": len(reader.pages),
        "used_ocr": False,
        "ocr_pages": [],
        "text_length": len(combined),
    }


def extract_text_from_pdf_bytes(file_bytes: bytes) -> Dict[str, Any]:
    if fitz is None:
        return _extract_pdf_text_with_reader(file_bytes)

    doc = fitz.open(stream=file_bytes, filetype="pdf")
    pages_text = []
    ocr_pages = []

    for page_number, page in enumerate(doc, start=1):
        text = (page.get_text("text") or "").strip()
        if len(text) >= 80 or RapidOCR is None:
            pages_text.append(text)
            continue

        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
        ocr_text = _ocr_image_bytes(pix.tobytes("png"))
        if ocr_text.strip():
            pages_text.append(ocr_text)
            ocr_pages.append(page_number)
        else:
            pages_text.append(text)

    combined = "\n".join([page_text for page_text in pages_text if page_text.strip()])
    return {
        "text": combined,
        "page_count": len(doc),
        "used_ocr": bool(ocr_pages),
        "ocr_pages": ocr_pages,
        "text_length": len(combined),
    }


def buscar_posicion_en_pdf(file_bytes: bytes, query: str) -> Optional[Dict[str, Any]]:
    """Busca en que pagina de un PDF aparece un fragmento de texto, usando el texto real
    del PDF (no el texto ya extraido por nuestro OCR), y devuelve tambien los rectangulos
    exactos donde aparece (en coordenadas PDF) para poder subrayarlo en el visor del
    frontend en vez de solo saltar a la pagina y dejar que el usuario lo busque a ojo."""
    if fitz is None or not query:
        return None

    consulta = re.sub(r"\s+", " ", query).strip()[:80]
    if not consulta:
        return None

    try:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
    except Exception:
        return None

    try:
        def _buscar_rects(needle: str):
            for page_number, page in enumerate(doc, start=1):
                try:
                    rects = page.search_for(needle, quads=False)
                except Exception:
                    rects = []
                if rects:
                    return page_number, page.rect.width, page.rect.height, [
                        [rect.x0, rect.y0, rect.x1, rect.y1] for rect in rects
                    ]
            return None

        # fitz.search_for tolera diferencias menores de espacio pero necesita el texto
        # tal cual aparece en el PDF; si el fragmento completo no calza (por diferencias
        # del OCR), se reintenta con un prefijo mas corto pero todavia distintivo.
        encontrado = _buscar_rects(consulta)
        consulta_corta = consulta[:40]
        if not encontrado and len(consulta_corta) >= 15:
            encontrado = _buscar_rects(consulta_corta)

        if encontrado:
            page_number, width, height, rects = encontrado
            return {"pagina": page_number, "rects": rects, "page_width": width, "page_height": height}

        # Ni siquiera search_for encontro nada (texto quizas partido en varias lineas
        # de forma que PyMuPDF no lo reconoce como una sola cadena); se cae al menos a
        # ubicar la pagina por substring sobre el texto normalizado, sin rectangulos.
        consulta_norm = re.sub(r"\s+", " ", consulta).lower()
        consulta_norm_corta = consulta_norm[:40]
        for page_number, page in enumerate(doc, start=1):
            texto_pagina = re.sub(r"\s+", " ", page.get_text("text") or "").lower()
            if consulta_norm in texto_pagina or (len(consulta_norm_corta) >= 15 and consulta_norm_corta in texto_pagina):
                return {"pagina": page_number, "rects": [], "page_width": page.rect.width, "page_height": page.rect.height}

        return None
    finally:
        doc.close()


def extract_text_from_file_bytes(file_bytes: bytes, filename: str = "") -> Dict[str, Any]:
    ext = os.path.splitext((filename or "").lower())[1]
    if ext in {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".webp"}:
        text = _ocr_image_bytes(file_bytes)
        return {
            "text": text,
            "page_count": 1,
            "used_ocr": True,
            "ocr_pages": [1] if text else [],
            "text_length": len(text),
        }

    if ext == ".pdf" or file_bytes[:4] == b"%PDF":
        return extract_text_from_pdf_bytes(file_bytes)

    try:
        text = file_bytes.decode("utf-8", errors="ignore")
    except Exception:
        text = ""

    return {
        "text": text,
        "page_count": 1,
        "used_ocr": False,
        "ocr_pages": [],
        "text_length": len(text),
    }



# Encabezado tolerante: "EXPERIENCIA No.1:", "EXPERIENCIA No 1", "EXPERIENCIA N°1 -",
# "EXPERIENCIA PROBABLE No.1:", etc. El OCR de distintas entidades/RUP nunca produce
# exactamente el mismo formato, asi que el separador entre "No" y el numero, y el que
# cierra el encabezado (":" o "-"), son opcionales.
_EXPERIENCIA_ENCABEZADO = r"(?:\*{0,3}\s*)?EXPERIENCIA(?:\s+PROBABLE)?\s*N[o°]?\.?\s*(\d+)\s*[:\-]?\s*"

EXPERIENCE_BLOCK_RE = re.compile(
    rf"{_EXPERIENCIA_ENCABEZADO}(.*?)(?={_EXPERIENCIA_ENCABEZADO}|\Z)",
    re.DOTALL | re.IGNORECASE,
)

# Codigo UNSPSC de 8 digitos en 4 grupos de 2 (segmento-familia-clase-producto). El
# separador entre grupos puede ser espacio, punto o guion (o ninguno, si viene pegado
# como "80101706"), y el cierre antes de la descripcion puede ser ":" o "-". Antes esto
# exigia grupos separados por espacio exacto y ":" fijo, por lo que un pliego o RUP
# formateado distinto (p.ej. "80.10.17.06 -" o "80101706:") no producia ningun codigo y
# la comparacion terminaba en "0 coincidencias" sin explicacion.
UNSPSC_RE = re.compile(
    r"(\d{2})[\s.\-]?(\d{2})[\s.\-]?(\d{2})[\s.\-]?(\d{2})\s*[:\-]\s*(.+)", re.IGNORECASE
)

# El RUP real (Certificado de Proponentes que expide cada Camara de Comercio) no trae
# bloques "EXPERIENCIA No.X:" ni codigos en formato "codigo: descripcion" -- cada contrato
# reportado empieza con "Numero consecutivo del reporte del contrato ejecutado: N" y sus
# codigos UNSPSC salen en una tabla "SEGMENTO FAMILIA CLASE PRODUCTO" (dos columnas por
# linea, sin ninguna descripcion). El "[uú�]" tolera tanto el acento correcto como el
# caracter de reemplazo "�" que deja un PDF con una codificacion de fuente rara al extraer
# texto (visto en certificados reales: "N�mero" en vez de "Número").
_CONSECUTIVO_LABEL = r"N[uú�]mero\s+consecutivo\s+del\s+reporte\s+del\s+contrato\s+ejecutado\s*:\s*(\d+)"

CONSECUTIVO_BLOCK_RE = re.compile(
    rf"{_CONSECUTIVO_LABEL}(.*?)(?={_CONSECUTIVO_LABEL}|\Z)",
    re.DOTALL | re.IGNORECASE,
)

_TABLA_CODIGOS_HEADER_RE = re.compile(r"SEGMENTO\s+FAMILIA\s+CLASE\s+PRODUCTO", re.IGNORECASE)
_TABLA_CODIGOS_CORTE_RE = re.compile(
    rf"{_CONSECUTIVO_LABEL}|P[aá�]gina\s*:", re.IGNORECASE
)


def _extraer_codigos_tabla_camara(bloque: str) -> List[Dict[str, Any]]:
    """Codigos UNSPSC de la tabla 'SEGMENTO FAMILIA CLASE PRODUCTO' del RUP real: sin
    descripcion, y con el campo PRODUCTO casi siempre en "00" (el RUP acredita a nivel de
    clase, no de producto especifico) -- por eso el cruce por clase/familia en
    buscar_coincidencias/comparar_codigos_pliego_rup es el que de verdad encuentra algo
    contra un codigo exacto exigido por el pliego."""
    header = _TABLA_CODIGOS_HEADER_RE.search(bloque)
    if not header:
        return []

    texto_tabla = bloque[header.end():]
    corte = _TABLA_CODIGOS_CORTE_RE.search(texto_tabla)
    if corte:
        texto_tabla = texto_tabla[: corte.start()]

    codigos: List[Dict[str, Any]] = []
    vistos = set()
    for match in re.finditer(r"(\d{2})\s+(\d{2})\s+(\d{2})\s+(\d{2})", texto_tabla):
        codigo_raw = "".join(match.groups())
        if codigo_raw in vistos:
            continue
        vistos.add(codigo_raw)
        codigos.append({"codigo": codigo_raw, "codigo_formateado": formatear_codigo(codigo_raw), "descripcion": ""})

    return codigos


def extraer_experiencias(texto: str) -> List[Dict[str, Any]]:
    texto = (texto or "").replace("\r", "\n")
    texto = re.sub(r"\n{2,}", "\n", texto)

    bloques = list(EXPERIENCE_BLOCK_RE.finditer(texto))
    formato_camara = not bloques
    if formato_camara:
        bloques = list(CONSECUTIVO_BLOCK_RE.finditer(texto))

    experiencias: List[Dict[str, Any]] = []
    for match in bloques:
        numero_exp = match.group(1).strip()
        bloque = match.group(2)

        consecutivo = numero_exp if formato_camara else ""
        contratista = ""
        contratante = ""
        valor_smmlv = ""

        if not formato_camara:
            consecutivo_match = re.search(r"N(?:ÚMERO|UMERO)\s+CONSECUTIVO\s+DEL\s+CONTRATO\s*:\s*(.+)", bloque, re.IGNORECASE)
            if consecutivo_match:
                consecutivo = consecutivo_match.group(1).strip().split("\n")[0].strip()

        contratista_match = re.search(r"NOMBRE\s+DEL\s+CONTRATISTA\s*:\s*(.+)", bloque, re.IGNORECASE)
        if contratista_match:
            contratista = contratista_match.group(1).strip().split("\n")[0].strip()

        contratante_match = re.search(r"NOMBRE\s+DEL\s+CONTRATANTE\s*:\s*(.+)", bloque, re.IGNORECASE)
        if contratante_match:
            contratante = contratante_match.group(1).strip().split("\n")[0].strip()

        # "VALOR CONTRATADO EN SMMLV:" (formato viejo asumido) o "VALOR DEL CONTRATO
        # EJECUTADO EXPRESADO EN SMMLV:" (RUP real) -- cualquier texto entre "VALOR" y
        # "SMMLV" antes de los dos puntos.
        valor_match = re.search(r"VALOR[^:\n]*SMMLV\s*:\s*(.+)", bloque, re.IGNORECASE)
        if valor_match:
            valor_smmlv = valor_match.group(1).strip().split("\n")[0].strip()

        codigos = []
        vistos_codigo = set()
        for code_match in UNSPSC_RE.finditer(bloque):
            codigo_raw = "".join(code_match.groups()[:4])
            if codigo_raw in vistos_codigo:
                continue
            vistos_codigo.add(codigo_raw)
            descripcion = code_match.group(5).strip()
            codigos.append(
                {
                    "codigo": codigo_raw,
                    "codigo_formateado": formatear_codigo(codigo_raw),
                    "descripcion": descripcion,
                }
            )

        for item in _extraer_codigos_tabla_camara(bloque):
            if item["codigo"] in vistos_codigo:
                continue
            vistos_codigo.add(item["codigo"])
            codigos.append(item)

        experiencias.append(
            {
                "experiencia_no": numero_exp,
                "consecutivo_contrato": consecutivo,
                "contratista": contratista,
                "contratante": contratante,
                "valor_smmlv": valor_smmlv,
                "codigos": codigos,
                "total_codigos": len(codigos),
            }
        )

    return experiencias


def buscar_coincidencias(experiencias: List[Dict[str, Any]], codigos_usuario: List[str]) -> List[Dict[str, Any]]:
    """Cruza los codigos que escribio el usuario contra los codigos de cada experiencia del
    RUP. Si no hay coincidencia exacta de codigo completo (8 digitos: segmento+familia+clase+
    producto), cae a clase (6 digitos) o familia (4 digitos) -- igual que el comparativo
    automatico pliego-vs-RUP -- para no perder coincidencias reales solo porque el producto
    exacto difiere dentro de la misma familia/clase UNSPSC."""
    codigos_usuario = [limpiar_codigo(c) for c in codigos_usuario if limpiar_codigo(c)]
    resultados: List[Dict[str, Any]] = []

    for experiencia in experiencias:
        codigos_experiencia = {item["codigo"]: item for item in experiencia["codigos"]}

        encontrados: List[str] = []
        detalle_coincidencias: List[Dict[str, Any]] = []

        for codigo in codigos_usuario:
            if codigo in codigos_experiencia:
                encontrados.append(codigo)
                detalle_coincidencias.append(
                    {
                        "codigo": formatear_codigo(codigo),
                        "descripcion": codigos_experiencia[codigo]["descripcion"],
                        "nivel_coincidencia": "codigo",
                    }
                )
                continue

            for nivel, largo in (("clase", 6), ("familia", 4)):
                prefijo = codigo[:largo]
                item_parcial = next(
                    (item for otro, item in codigos_experiencia.items() if otro[:largo] == prefijo), None
                )
                if item_parcial:
                    encontrados.append(codigo)
                    detalle_coincidencias.append(
                        {
                            "codigo": formatear_codigo(codigo),
                            "descripcion": item_parcial["descripcion"],
                            "nivel_coincidencia": nivel,
                        }
                    )
                    break

        if not encontrados:
            continue

        resultados.append(
            {
                "experiencia_no": experiencia["experiencia_no"],
                "consecutivo_contrato": experiencia["consecutivo_contrato"],
                "contratante": experiencia["contratante"],
                "contratista": experiencia["contratista"],
                "valor_smmlv": experiencia["valor_smmlv"],
                "codigos_encontrados": [formatear_codigo(codigo) for codigo in encontrados],
                "cantidad_coincidencias": len(encontrados),
                "detalle_coincidencias": detalle_coincidencias,
            }
        )

    return sorted(
        resultados,
        key=lambda item: (-item["cantidad_coincidencias"], int(item["experiencia_no"]) if str(item["experiencia_no"]).isdigit() else 0),
    )


_TABLA_PLIEGO_HEADER_RE = re.compile(
    r"C.digo\s*Segmento\s*C.digo\s*Familia\s*C.digo\s*Clase\s*Descripci.n\s*del\s*producto",
    re.IGNORECASE,
)
_TABLA_PLIEGO_CORTE_RE = re.compile(r"Proyecto de Pliego de Condiciones|P.gina\s*\d+\s*de\s*\d+", re.IGNORECASE)
_TABLA_PLIEGO_FILA_RE = re.compile(r"(\d{6,8})\s+(\d{6,8})\s+(\d{8})\s+")


def _extraer_codigos_tabla_pliego(texto: str) -> List[Dict[str, Any]]:
    """Tabla real de clasificacion UNSPSC que usan los pliegos: encabezado 'Codigo Segmento
    / Codigo Familia / Codigo Clase / Descripcion del producto' y, por cada renglon, tres
    codigos seguidos (uno por linea) mas la descripcion del producto. El codigo de clase
    (tercero) es el unico que llega confiable a 8 digitos completos -- el de segmento a
    veces pierde un digito en la extraccion/OCR (p.ej. '9000000' en vez de '90000000') -- y
    ya incluye segmento+familia en sus primeros 4 digitos, asi que es el que se usa."""
    if not texto:
        return []

    header = _TABLA_PLIEGO_HEADER_RE.search(texto)
    if not header:
        return []

    texto_tabla = texto[header.end():]
    corte = _TABLA_PLIEGO_CORTE_RE.search(texto_tabla)
    if corte:
        texto_tabla = texto_tabla[: corte.start()]

    filas = list(_TABLA_PLIEGO_FILA_RE.finditer(texto_tabla))
    resultados: List[Dict[str, Any]] = []
    vistos = set()
    for idx, fila in enumerate(filas):
        codigo_raw = fila.group(3)
        if codigo_raw in vistos:
            continue
        vistos.add(codigo_raw)
        inicio_desc = fila.end()
        fin_desc = filas[idx + 1].start() if idx + 1 < len(filas) else len(texto_tabla)
        descripcion = re.sub(r"\s+", " ", texto_tabla[inicio_desc:fin_desc]).strip()[:200]
        resultados.append(
            {"codigo": codigo_raw, "codigo_formateado": formatear_codigo(codigo_raw), "descripcion": descripcion}
        )

    return resultados


def extraer_codigos_pliego(texto: str) -> List[Dict[str, Any]]:
    """Codigos UNSPSC exigidos por el pliego. Se intentan dos formatos reales: la tabla
    'Codigo Segmento / Familia / Clase / Descripcion del producto' que usan la mayoria de
    pliegos (ver _extraer_codigos_tabla_pliego), y el patron suelto 'XX XX XX XX: descripcion'
    que usan otros como prosa."""
    if not texto:
        return []

    resultados: List[Dict[str, Any]] = []
    vistos = set()

    for item in _extraer_codigos_tabla_pliego(texto):
        if item["codigo"] in vistos:
            continue
        vistos.add(item["codigo"])
        resultados.append(item)

    for match in UNSPSC_RE.finditer(texto):
        codigo_raw = "".join(match.groups()[:4])
        if codigo_raw in vistos:
            continue
        descripcion = match.group(5).strip().split("\n")[0][:200]
        # El patron "XX XX XX XX: algo" tambien calza por coincidencia con codigos
        # presupuestales, radicados judiciales, etc. que no tienen ninguna descripcion real
        # al lado (solo mas numeros/puntuacion); exigir letras de verdad filtra ese ruido.
        if not re.search(r"[a-zA-ZÁÉÍÓÚÑáéíóúñ]{3,}", descripcion):
            continue
        vistos.add(codigo_raw)
        resultados.append(
            {
                "codigo": codigo_raw,
                "codigo_formateado": formatear_codigo(codigo_raw),
                "descripcion": descripcion,
            }
        )
    return resultados


_STOPWORDS_OBJETO = {
    "para", "con", "por", "los", "las", "del", "que", "una", "uno", "sus", "este", "esta",
    "sobre", "entre", "como", "cada", "todo", "toda", "todos", "todas", "mas", "sin", "segun",
}


def _palabras_significativas(texto: Optional[str]) -> set:
    palabras = re.findall(r"[a-zA-ZÁÉÍÓÚÑÜáéíóúñü]{4,}", texto or "")
    return {p.lower() for p in palabras if p.lower() not in _STOPWORDS_OBJETO}


def buscar_coincidencias_objeto_contractual(
    objeto_contrato: Optional[str], experiencias_rup: List[Dict[str, Any]], minimo: int = 2
) -> List[Dict[str, Any]]:
    """Complementa la busqueda por codigo UNSPSC: cuando el pliego describe el objeto en
    prosa (sin codigos claros), compara las palabras clave del objeto del contrato contra
    las descripciones de cada experiencia del RUP."""
    palabras_objeto = _palabras_significativas(objeto_contrato)
    if not palabras_objeto:
        return []

    resultados = []
    for experiencia in experiencias_rup:
        descripciones = " ".join(codigo["descripcion"] for codigo in experiencia["codigos"])
        palabras_comunes = palabras_objeto & _palabras_significativas(descripciones)
        if len(palabras_comunes) >= minimo:
            resultados.append(
                {
                    "experiencia_no": experiencia["experiencia_no"],
                    "contratante": experiencia["contratante"],
                    "contratista": experiencia["contratista"],
                    "palabras_comunes": sorted(palabras_comunes),
                    "cantidad_palabras_comunes": len(palabras_comunes),
                }
            )

    return sorted(resultados, key=lambda row: -row["cantidad_palabras_comunes"])[:10]


def comparar_codigos_pliego_rup(
    pliego_texto: Optional[str], rup_texto: Optional[str], objeto_contrato: Optional[str] = None
) -> Optional[Dict[str, Any]]:
    """Compara los codigos UNSPSC exigidos por el pliego contra los que aparecen en las
    experiencias del RUP, para saber cuales ya se pueden acreditar (herramienta integrada
    de coincidencias RUP x UNSPSC del pliego). Si no hay coincidencia exacta de codigo
    (8 digitos: segmento+familia+clase+producto), intenta por clase (6 digitos) o familia
    (4 digitos); y complementa con coincidencias por objeto contractual."""
    if not pliego_texto or not rup_texto:
        return None

    codigos_pliego = extraer_codigos_pliego(pliego_texto)
    experiencias_rup = extraer_experiencias(rup_texto)

    rup_por_codigo: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for experiencia in experiencias_rup:
        for codigo in experiencia["codigos"]:
            rup_por_codigo[codigo["codigo"]].append(
                {
                    "experiencia_no": experiencia["experiencia_no"],
                    "contratante": experiencia["contratante"],
                    "contratista": experiencia["contratista"],
                }
            )

    def _coincidencia(codigo: str) -> Optional[Dict[str, Any]]:
        if codigo in rup_por_codigo:
            return {"nivel": "codigo", "experiencias": rup_por_codigo[codigo]}

        for nivel, largo in (("clase", 6), ("familia", 4)):
            prefijo = codigo[:largo]
            experiencias = [
                exp for otro, exps in rup_por_codigo.items() if otro[:largo] == prefijo for exp in exps
            ]
            if experiencias:
                return {"nivel": nivel, "experiencias": experiencias}

        return None

    comunes = []
    faltantes = []
    for item in codigos_pliego:
        entry = {
            "codigo": item["codigo"],
            "codigo_formateado": item["codigo_formateado"],
            "descripcion_pliego": item["descripcion"],
        }
        coincidencia = _coincidencia(item["codigo"])
        if coincidencia:
            comunes.append({**entry, "experiencias_rup": coincidencia["experiencias"], "nivel_coincidencia": coincidencia["nivel"]})
        else:
            faltantes.append(entry)

    # Diagnostico para cuando "0 coincidencias" en realidad significa que no se pudo
    # extraer nada de uno de los dos documentos (formato distinto al esperado, OCR
    # ilegible, etc.), en vez de que de verdad no haya cruce entre pliego y RUP.
    diagnostico = None
    if not codigos_pliego:
        diagnostico = "sin_codigos_pliego"
    elif not experiencias_rup:
        diagnostico = "sin_experiencias_rup"
    elif not rup_por_codigo:
        diagnostico = "experiencias_rup_sin_codigos"

    return {
        "codigos_pliego_total": len(codigos_pliego),
        "codigos_rup_total": len(rup_por_codigo),
        "experiencias_rup_total": len(experiencias_rup),
        "codigos_comunes": sorted(comunes, key=lambda row: row["codigo"]),
        "codigos_faltantes": sorted(faltantes, key=lambda row: row["codigo"]),
        "coincidencias_objeto": buscar_coincidencias_objeto_contractual(objeto_contrato, experiencias_rup),
        "diagnostico": diagnostico,
    }


def tabla_codigos(experiencias: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for experiencia in experiencias:
        for codigo in experiencia["codigos"]:
            rows.append(
                {
                    "experiencia_no": experiencia["experiencia_no"],
                    "consecutivo_contrato": experiencia["consecutivo_contrato"],
                    "contratante": experiencia["contratante"],
                    "contratista": experiencia["contratista"],
                    "codigo": codigo["codigo_formateado"],
                    "descripcion": codigo["descripcion"],
                }
            )
    return rows


def _extract_numbers(texto: str) -> List[float]:
    tokens = re.findall(r"(?<!\w)(?:\d{1,3}(?:[.,]\d{3})+|\d+)(?:[.,]\d+)?", texto or "")
    numbers: List[float] = []
    for token in tokens:
        normalized = token.replace(".", "").replace(",", ".")
        try:
            numbers.append(float(normalized))
        except ValueError:
            continue
    return numbers


REQUISITO_KEYWORD_RE = re.compile(
    r"(?i)(el proponente deber|la propuesta deber|el oferente deber|se requiere|es obligatorio|"
    r"debe(?:r[aá])? acreditar|debe(?:r[aá])? presentar|debe(?:r[aá])? anexar|debe(?:r[aá])? cumplir|"
    r"debe(?:r[aá])? garantizar)"
)


def extraer_requisitos_sugeridos(texto: str, limite: int = 15) -> List[str]:
    """Heuristica: lineas del pliego que probablemente describen un requisito exigido."""
    if not texto:
        return []

    sugerencias: List[str] = []
    vistos = set()

    for linea in re.split(r"[\n\r]+", texto):
        limpio = re.sub(r"\s+", " ", linea).strip(" .-*•")
        if not (15 <= len(limpio) <= 220):
            continue
        if not REQUISITO_KEYWORD_RE.search(limpio):
            continue

        firma = normalize_text(limpio)[:120]
        if firma in vistos:
            continue
        vistos.add(firma)
        sugerencias.append(limpio)
        if len(sugerencias) >= limite:
            break

    return sugerencias


_FACTOR_DESEMPATE_RE = re.compile(
    r"factor(?:es)?\s+de\s+desempate|criterio(?:s)?\s+de\s+desempate|en\s+caso\s+de\s+empate",
    re.IGNORECASE,
)


def _extraer_parrafo(texto: str, inicio: int, fin: int, ventana: int = 500) -> str:
    """Devuelve el parrafo completo alrededor de una coincidencia. Si el documento no usa
    doble salto de linea entre parrafos (comun en PDFs escaneados/aplanados por OCR), cae a una
    ventana de caracteres fija alrededor del match."""
    inicio_parrafo = texto.rfind("\n\n", 0, inicio)
    inicio_parrafo = 0 if inicio_parrafo == -1 else inicio_parrafo + 2
    fin_parrafo = texto.find("\n\n", fin)
    fin_parrafo = len(texto) if fin_parrafo == -1 else fin_parrafo
    parrafo = texto[inicio_parrafo:fin_parrafo].strip()

    if len(parrafo) < 40:
        desde = max(0, inicio - ventana)
        hasta = min(len(texto), fin + ventana)
        parrafo = texto[desde:hasta].strip()

    return re.sub(r"\s+", " ", parrafo)


def extraer_factor_desempate(texto: str) -> List[Dict[str, Any]]:
    """Busca las clausulas de desempate del pliego (Ley 2069/2020, Decreto 1082) y devuelve el
    parrafo completo donde aparecen, no solo si existen o no."""
    if not texto:
        return []

    hallazgos = []
    vistos = set()
    for match in _FACTOR_DESEMPATE_RE.finditer(texto):
        parrafo = _extraer_parrafo(texto, match.start(), match.end())
        firma = parrafo[:150].lower()
        if firma in vistos:
            continue
        vistos.add(firma)
        hallazgos.append({"coincidencia": match.group(0), "texto": parrafo})
        if len(hallazgos) >= 5:
            break

    return hallazgos


def analyze_text(texto: str, codigos_usuario: Optional[List[str]] = None) -> Dict[str, Any]:
    experiencias = extraer_experiencias(texto)
    codigos_busqueda = [limpiar_codigo(c) for c in (codigos_usuario or []) if limpiar_codigo(c)]
    coincidencias = buscar_coincidencias(experiencias, codigos_busqueda) if codigos_busqueda else []
    codigos_extraidos = tabla_codigos(experiencias)

    numeros = _extract_numbers(texto)
    porcentajes = re.findall(r"\b\d+(?:[.,]\d+)?\s*%", texto or "")
    smmlv = re.findall(r"\b\d+(?:[.,]\d+)?\s*SMMLV\b", texto or "", re.IGNORECASE)
    monedas = re.findall(r"\$\s*[\d\.,]+", texto or "")

    resumen = {
        "experiencias_detectadas": len(experiencias),
        "codigos_extraidos": len(codigos_extraidos),
        "coincidencias_detectadas": len(coincidencias),
        "codigos_busqueda": [formatear_codigo(codigo) for codigo in codigos_busqueda],
        "numeros_detectados": len(numeros),
        "suma_numeros": round(sum(numeros), 2) if numeros else 0,
        "promedio_numeros": round(sum(numeros) / len(numeros), 2) if numeros else 0,
        "maximo_numero": max(numeros) if numeros else 0,
        "minimo_numero": min(numeros) if numeros else 0,
        "porcentajes_detectados": len(porcentajes),
        "smmlv_menciones": len(smmlv),
        "valores_monetarios": len(monedas),
    }

    return {
        "experiencias": experiencias,
        "coincidencias": coincidencias,
        "codigos_extraidos": codigos_extraidos,
        "resumen": resumen,
        "texto_preview": (texto or "")[:5000],
        "texto_completo": texto or "",
        "requisitos_sugeridos": extraer_requisitos_sugeridos(texto),
        "factor_desempate": extraer_factor_desempate(texto),
    }


def analyze_file_bytes(file_bytes: bytes, filename: str = "", codigos_usuario: Optional[List[str]] = None) -> Dict[str, Any]:
    extracted = extract_text_from_file_bytes(file_bytes, filename)
    analysis = analyze_text(extracted["text"], codigos_usuario)
    return {
        "archivo": {
            "nombre": filename,
            "page_count": extracted["page_count"],
            "used_ocr": extracted["used_ocr"],
            "ocr_pages": extracted["ocr_pages"],
            "text_length": extracted["text_length"],
        },
        **analysis,
    }


def build_analysis_xlsx_bytes(analysis: Dict[str, Any]) -> bytes:
    if Workbook is None:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="La exportacion a Excel requiere openpyxl. Instala openpyxl en el backend o usa CSV.",
        )

    workbook = Workbook()

    ws = workbook.active
    ws.title = "Coincidencias"
    coincidencias = analysis.get("coincidencias", [])
    if coincidencias:
        headers = [
            "experiencia_no",
            "consecutivo_contrato",
            "contratante",
            "contratista",
            "valor_smmlv",
            "cantidad_coincidencias",
            "codigos_encontrados",
            "detalle_coincidencias",
        ]
        ws.append(headers)
        for row in coincidencias:
            ws.append(
                [
                    row.get("experiencia_no"),
                    row.get("consecutivo_contrato"),
                    row.get("contratante"),
                    row.get("contratista"),
                    row.get("valor_smmlv"),
                    row.get("cantidad_coincidencias"),
                    ", ".join(row.get("codigos_encontrados", [])),
                    " | ".join(
                        f"{item['codigo']} - {item['descripcion']}" for item in row.get("detalle_coincidencias", [])
                    ),
                ]
            )
    else:
        ws.append(["Sin coincidencias"])

    ws_summary = workbook.create_sheet("Resumen")
    for key, value in (analysis.get("resumen") or {}).items():
        if isinstance(value, list):
            value = ", ".join(map(str, value))
        ws_summary.append([key, value])

    ws_exp = workbook.create_sheet("Experiencias")
    ws_exp.append(["experiencia_no", "consecutivo_contrato", "contratante", "contratista", "valor_smmlv", "total_codigos"])
    for exp in analysis.get("experiencias", []):
        ws_exp.append(
            [
                exp.get("experiencia_no"),
                exp.get("consecutivo_contrato"),
                exp.get("contratante"),
                exp.get("contratista"),
                exp.get("valor_smmlv"),
                exp.get("total_codigos"),
            ]
        )

    ws_codes = workbook.create_sheet("Codigos")
    ws_codes.append(["experiencia_no", "consecutivo_contrato", "contratante", "contratista", "codigo", "descripcion"])
    for row in analysis.get("codigos_extraidos", []):
        ws_codes.append(
            [
                row.get("experiencia_no"),
                row.get("consecutivo_contrato"),
                row.get("contratante"),
                row.get("contratista"),
                row.get("codigo"),
                row.get("descripcion"),
            ]
        )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.read()


def build_analysis_csv_text(analysis: Dict[str, Any]) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "experiencia_no",
            "consecutivo_contrato",
            "contratante",
            "contratista",
            "valor_smmlv",
            "cantidad_coincidencias",
            "codigos_encontrados",
            "detalle_coincidencias",
        ]
    )

    for row in analysis.get("coincidencias", []):
        writer.writerow(
            [
                row.get("experiencia_no"),
                row.get("consecutivo_contrato"),
                row.get("contratante"),
                row.get("contratista"),
                row.get("valor_smmlv"),
                row.get("cantidad_coincidencias"),
                ", ".join(row.get("codigos_encontrados", [])),
                " | ".join(f"{item['codigo']} - {item['descripcion']}" for item in row.get("detalle_coincidencias", [])),
            ]
        )

    return output.getvalue()


def build_checklist_pdf_bytes(licitacion, items: List[Dict[str, Any]], resumen: Dict[str, Any]) -> bytes:
    """Genera un PDF simple con el estado del checklist (documentos obligatorios y personalizados)."""
    if fitz is None:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="La generacion de PDF requiere PyMuPDF instalado en el backend.",
        )

    doc = fitz.open()
    page = doc.new_page()
    margin = 50
    state = {"y": 60, "page": page}

    def draw(text: str, size: float = 11, font: str = "helv", color=(0, 0, 0), gap: float = 8):
        if state["y"] > state["page"].rect.height - 60:
            state["page"] = doc.new_page()
            state["y"] = 60
        state["page"].insert_text((margin, state["y"]), text, fontsize=size, fontname=font, color=color)
        state["y"] += size + gap

    draw("CHECKLIST DE DOCUMENTOS", size=16, font="hebo")
    draw(f"Proceso: {licitacion.numero_secop or 'Sin numero'}", size=11, font="hebo")
    draw(f"Entidad: {licitacion.entidad_contratante or 'Sin definir'}", size=11)
    draw(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", size=9, color=(0.45, 0.45, 0.45))
    state["y"] += 4
    draw(
        f"Cobertura obligatoria: {resumen.get('cobertura_porcentaje', 0)}% "
        f"({resumen.get('obligatorios_cumplidos', 0)}/{resumen.get('obligatorios_total', 0)} cumplidos)",
        size=12,
        font="hebo",
    )
    state["y"] += 8

    for item in items:
        cumple = bool(item.get("cumple"))
        etiqueta = "CUMPLE" if cumple else "PENDIENTE"
        color = (0, 0.5, 0) if cumple else (0.75, 0, 0)
        obligatorio = " (obligatorio)" if item.get("obligatorio") else " (opcional)"
        draw(f"[{etiqueta}] {item.get('nombre', '')}{obligatorio}", size=11, color=color)
        detalle = item.get("documento_nombre") or item.get("descripcion")
        if detalle:
            draw(f"    {str(detalle)[:110]}", size=9, color=(0.4, 0.4, 0.4), gap=10)

    pdf_bytes = doc.tobytes()
    doc.close()
    return pdf_bytes


def build_comparativo_pdf_bytes(licitacion, comparativo: Dict[str, Any]) -> bytes:
    """Genera un PDF con el cruce de codigos UNSPSC (pliego vs RUP) y las coincidencias
    por objeto contractual, para dejar constancia del resultado del analisis."""
    if fitz is None:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="La generacion de PDF requiere PyMuPDF instalado en el backend.",
        )

    NIVEL_ETIQUETA = {"codigo": "", "clase": " (coincide por clase)", "familia": " (coincide por familia)"}

    doc = fitz.open()
    page = doc.new_page()
    margin = 50
    state = {"y": 60, "page": page}

    def draw(text: str, size: float = 11, font: str = "helv", color=(0, 0, 0), gap: float = 8):
        if state["y"] > state["page"].rect.height - 60:
            state["page"] = doc.new_page()
            state["y"] = 60
        state["page"].insert_text((margin, state["y"]), text, fontsize=size, fontname=font, color=color)
        state["y"] += size + gap

    draw("COINCIDENCIAS UNSPSC: PLIEGO x RUP", size=16, font="hebo")
    draw(f"Proceso: {licitacion.numero_secop or 'Sin numero'}", size=11, font="hebo")
    draw(f"Entidad: {licitacion.entidad_contratante or 'Sin definir'}", size=11)
    draw(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", size=9, color=(0.45, 0.45, 0.45))
    state["y"] += 4
    draw(
        f"Coincidencias: {len(comparativo['codigos_comunes'])}/{comparativo['codigos_pliego_total']} "
        "codigos del pliego ya acredita el RUP",
        size=12,
        font="hebo",
    )
    state["y"] += 8

    draw("CODIGOS EN COMUN", size=12, font="hebo")
    if not comparativo["codigos_comunes"]:
        draw("Ningun codigo del pliego coincide todavia con el RUP.", size=10, color=(0.6, 0.2, 0.2))
    for item in comparativo["codigos_comunes"]:
        etiqueta_nivel = NIVEL_ETIQUETA.get(item.get("nivel_coincidencia"), "")
        draw(f"[OK] {item['codigo_formateado']}{etiqueta_nivel} - {item['descripcion_pliego'][:90]}", size=10, color=(0, 0.5, 0))
        for exp in item["experiencias_rup"][:3]:
            draw(f"      Exp. #{exp['experiencia_no']} - {exp['contratante']}", size=9, color=(0.4, 0.4, 0.4), gap=4)
    state["y"] += 6

    if comparativo["codigos_faltantes"]:
        draw("CODIGOS QUE AUN NO ACREDITA EL RUP", size=12, font="hebo")
        for item in comparativo["codigos_faltantes"]:
            draw(f"[ ] {item['codigo_formateado']} - {item['descripcion_pliego'][:90]}", size=10, color=(0.75, 0, 0))
        state["y"] += 6

    if comparativo.get("coincidencias_objeto"):
        draw("COINCIDENCIAS POR OBJETO CONTRACTUAL", size=12, font="hebo")
        for item in comparativo["coincidencias_objeto"]:
            draw(
                f"Exp. #{item['experiencia_no']} - {item['contratante']} "
                f"({item['cantidad_palabras_comunes']} palabras en comun: {', '.join(item['palabras_comunes'][:6])})",
                size=9.5,
            )

    pdf_bytes = doc.tobytes()
    doc.close()
    return pdf_bytes


def serialize_empresa(empresa: Empresa) -> Dict[str, Any]:
    return {
        "id": str(empresa.id),
        "grupo_empresarial_id": str(empresa.grupo_empresarial_id) if empresa.grupo_empresarial_id else None,
        "nombre": empresa.nombre,
        "nit": empresa.nit,
        "direccion": empresa.direccion,
        "telefono": empresa.telefono,
        "email": empresa.email,
        "sitio_web": empresa.sitio_web,
        "logo_url": empresa.logo_url,
        "activo": empresa.activo,
        "created_at": empresa.created_at.isoformat() if empresa.created_at else None,
        "updated_at": empresa.updated_at.isoformat() if empresa.updated_at else None,
    }


def serialize_licitacion(licitacion: Licitacion) -> Dict[str, Any]:
    cuantia = float(licitacion.cuantia) if licitacion.cuantia is not None else None
    dias_restantes = None
    if licitacion.fecha_cierre:
        dias_restantes = (licitacion.fecha_cierre.date() - datetime.now().date()).days

    return {
        "id": str(licitacion.id),
        "empresa_id": str(licitacion.empresa_id),
        "numero_secop": licitacion.numero_secop,
        "url_secop": licitacion.url_secop,
        "entidad_contratante": licitacion.entidad_contratante,
        "nit_entidad": licitacion.nit_entidad,
        "objeto_contrato": licitacion.objeto_contrato,
        "cuantia": cuantia,
        "estado": licitacion.estado,
        "fecha_publicacion": licitacion.fecha_publicacion.isoformat() if licitacion.fecha_publicacion else None,
        "fecha_apertura": licitacion.fecha_apertura.isoformat() if licitacion.fecha_apertura else None,
        "fecha_cierre": licitacion.fecha_cierre.isoformat() if licitacion.fecha_cierre else None,
        "fecha_subsanacion": licitacion.fecha_subsanacion.isoformat() if licitacion.fecha_subsanacion else None,
        "fecha_adjudicacion": licitacion.fecha_adjudicacion.isoformat() if licitacion.fecha_adjudicacion else None,
        "fecha_visita_obra": licitacion.fecha_visita_obra.isoformat() if licitacion.fecha_visita_obra else None,
        "fecha_consultas": licitacion.fecha_consultas.isoformat() if licitacion.fecha_consultas else None,
        "fecha_cierre_dudas": licitacion.fecha_cierre_dudas.isoformat() if licitacion.fecha_cierre_dudas else None,
        "fecha_evaluacion": licitacion.fecha_evaluacion.isoformat() if licitacion.fecha_evaluacion else None,
        "pliego_url": licitacion.pliego_url,
        "pliego_texto": licitacion.pliego_texto,
        "rup_url": licitacion.rup_url,
        "rup_texto": licitacion.rup_texto,
        "notas": licitacion.notas,
        "usuario_creador": str(licitacion.usuario_creador) if licitacion.usuario_creador else None,
        "created_at": licitacion.created_at.isoformat() if licitacion.created_at else None,
        "updated_at": licitacion.updated_at.isoformat() if licitacion.updated_at else None,
        "dias_restantes": dias_restantes,
    }


def serialize_documento(documento: Documento) -> Dict[str, Any]:
    return {
        "id": str(documento.id),
        "carpeta_id": str(documento.carpeta_id) if documento.carpeta_id else None,
        "empresa_id": str(documento.empresa_id),
        "nombre": documento.nombre,
        "nombre_original": documento.nombre_original,
        "tipo_documento": documento.tipo_documento,
        "descripcion": documento.descripcion,
        "ruta_archivo": documento.ruta_archivo,
        "tamanio_bytes": documento.tamanio_bytes,
        "formato": documento.formato,
        "version": documento.version,
        "vigente": documento.vigente,
        "fecha_vencimiento": documento.fecha_vencimiento.isoformat() if documento.fecha_vencimiento else None,
        "tags": documento.tags,
        "meta_data": documento.meta_data or {},
        "usuario_subida": str(documento.usuario_subida) if documento.usuario_subida else None,
        "created_at": documento.created_at.isoformat() if documento.created_at else None,
        "updated_at": documento.updated_at.isoformat() if documento.updated_at else None,
    }


def serialize_carpeta(carpeta: Carpeta) -> Dict[str, Any]:
    return {
        "id": str(carpeta.id),
        "empresa_id": str(carpeta.empresa_id),
        "carpeta_padre_id": str(carpeta.carpeta_padre_id) if carpeta.carpeta_padre_id else None,
        "nombre": carpeta.nombre,
        "descripcion": carpeta.descripcion,
        "icono": carpeta.icono,
        "color": carpeta.color,
        "created_at": carpeta.created_at.isoformat() if carpeta.created_at else None,
        "updated_at": carpeta.updated_at.isoformat() if carpeta.updated_at else None,
    }


def build_folder_tree(folders: List[Carpeta], documents: List[Documento]) -> Dict[str, Any]:
    folders_by_parent: Dict[Optional[str], List[Carpeta]] = defaultdict(list)
    docs_by_folder: Dict[Optional[str], List[Documento]] = defaultdict(list)

    for folder in folders:
        parent_id = str(folder.carpeta_padre_id) if folder.carpeta_padre_id else None
        folders_by_parent[parent_id].append(folder)

    for document in documents:
        folder_id = str(document.carpeta_id) if document.carpeta_id else None
        docs_by_folder[folder_id].append(document)

    def build_node(folder: Carpeta, level: int, prefix: str) -> Dict[str, Any]:
        node_path = f"{prefix}/{folder.nombre}" if prefix else folder.nombre
        folder_id = str(folder.id)
        child_nodes = [build_node(child, level + 1, node_path) for child in sorted(folders_by_parent.get(folder_id, []), key=lambda item: item.nombre.lower())]
        folder_docs = [serialize_documento(doc) for doc in sorted(docs_by_folder.get(folder_id, []), key=lambda item: item.nombre.lower())]

        return {
            **serialize_carpeta(folder),
            "nivel": level,
            "ruta": node_path,
            "documentos": folder_docs,
            "hijos": child_nodes,
            "total_documentos": len(folder_docs),
            "total_hijos": len(child_nodes),
        }

    arbol = [build_node(folder, 0, "") for folder in sorted(folders_by_parent.get(None, []), key=lambda item: item.nombre.lower())]
    documentos_raiz = [serialize_documento(doc) for doc in sorted(docs_by_folder.get(None, []), key=lambda item: item.nombre.lower())]

    return {
        "arbol": arbol,
        "documentos_raiz": documentos_raiz,
    }


def _document_signature(documento: Documento) -> str:
    parts = [
        documento.nombre,
        documento.nombre_original,
        documento.tipo_documento,
        documento.descripcion,
        documento.tags,
        json.dumps(documento.meta_data or {}, ensure_ascii=False) if documento.meta_data else "",
    ]
    return normalize_text(" ".join(filter(None, parts)))


def _apply_estado_manual(item: Dict[str, Any], estado, documentos_por_id: Dict[str, Documento]) -> Dict[str, Any]:
    """Aplica el estado manual (checkbox) del checklist sobre un item base.
    "cumplido" es SIEMPRE decision manual de una persona; el documento adjunto es opcional."""
    if estado is None:
        item.update(
            {
                "cumple": False,
                "documento_id": None,
                "documento_nombre": None,
                "validado_por": None,
                "validado_por_nombre": None,
                "validado_en": None,
                "requiere_subsanacion": False,
                "notas_subsanacion": None,
            }
        )
        return item

    documento_adjunto = documentos_por_id.get(str(estado.documento_id)) if estado.documento_id else None

    item.update(
        {
            "cumple": bool(estado.cumplido),
            "documento_id": str(estado.documento_id) if estado.documento_id else None,
            "documento_nombre": documento_adjunto.nombre if documento_adjunto else None,
            "validado_por": str(estado.validado_por) if estado.validado_por else None,
            "validado_por_nombre": None,  # se completa en build_checklist_items (requiere lookup de usuarios)
            "validado_en": estado.validado_en.isoformat() if estado.validado_en else None,
            "requiere_subsanacion": bool(estado.requiere_subsanacion),
            "notas_subsanacion": estado.notas_subsanacion,
        }
    )
    return item


def build_required_documents_status(
    estados_map: Dict[str, Any], documentos_por_id: Dict[str, Documento], excluidos: Optional[List[str]] = None
) -> List[Dict[str, Any]]:
    excluidos_set = set(excluidos or [])

    result = []
    for rule in DEFAULT_REQUIRED_DOCUMENTS:
        if rule["key"] in excluidos_set:
            continue

        item = {
            "key": rule["key"],
            "nombre": rule["nombre"],
            "descripcion": rule["descripcion"],
            "categoria": rule["categoria"],
            "obligatorio": rule["obligatorio"],
            "personalizado": False,
            "requisito_id": None,
            "excluible": True,
        }
        result.append(_apply_estado_manual(item, estados_map.get(rule["key"]), documentos_por_id))

    return result


def build_checklist_items(
    licitacion_id: str, documents: List[Documento], db: Session, excluidos: Optional[List[str]] = None
) -> List[Dict[str, Any]]:
    """Checklist base (DEFAULT_REQUIRED_DOCUMENTS, menos los excluidos) + requisitos personalizados.
    El estado "cumplido" es siempre manual (checkbox), guardado en ChecklistEstado; subir un
    documento es opcional y no marca nada por si solo."""
    from src.controllers.checklist_estado_controller import ChecklistEstadoController

    estados_map = ChecklistEstadoController.get_estados_map(licitacion_id, db)
    documentos_por_id = {str(documento.id): documento for documento in documents}

    items = build_required_documents_status(estados_map, documentos_por_id, excluidos)

    custom_requisitos = (
        db.query(RequisitoChecklist)
        .filter(RequisitoChecklist.licitacion_id == licitacion_id, RequisitoChecklist.tipo == "global")
        .order_by(RequisitoChecklist.orden, RequisitoChecklist.created_at)
        .all()
    )

    for requisito in custom_requisitos:
        item_key = f"custom:{requisito.id}"
        item = {
            "key": item_key,
            "requisito_id": str(requisito.id),
            "nombre": requisito.nombre,
            "descripcion": requisito.descripcion_original or "Requisito agregado para esta licitacion",
            "categoria": "personalizado",
            "obligatorio": requisito.obligatorio if requisito.obligatorio is not None else True,
            "personalizado": True,
        }
        items.append(_apply_estado_manual(item, estados_map.get(item_key), documentos_por_id))

    usuario_ids = {item["validado_por"] for item in items if item.get("validado_por")}
    if usuario_ids:
        import uuid as _uuid

        nombres = ChecklistEstadoController.nombres_usuarios(
            (_uuid.UUID(uid) for uid in usuario_ids), db
        )
        for item in items:
            if item.get("validado_por"):
                item["validado_por_nombre"] = nombres.get(_uuid.UUID(item["validado_por"]))

    return items


def build_checklist_actividad(licitacion_id: str, db: Session, limit: int = 15) -> List[Dict[str, Any]]:
    """Actividad reciente del checklist (quien marco cada item, cuando), con el nombre del
    item resuelto contra el checklist actual (default + personalizado)."""
    from src.controllers.checklist_estado_controller import ChecklistEstadoController

    licitacion_row = db.query(Licitacion).filter(Licitacion.id == licitacion_id).first()
    excluidos = licitacion_row.checklist_excluidos if licitacion_row else None

    nombres_por_key = {rule["key"]: rule["nombre"] for rule in DEFAULT_REQUIRED_DOCUMENTS}
    excluidos_set = set(excluidos or [])
    nombres_por_key = {key: nombre for key, nombre in nombres_por_key.items() if key not in excluidos_set}

    custom_requisitos = (
        db.query(RequisitoChecklist)
        .filter(RequisitoChecklist.licitacion_id == licitacion_id, RequisitoChecklist.tipo == "global")
        .all()
    )
    for requisito in custom_requisitos:
        nombres_por_key[f"custom:{requisito.id}"] = requisito.nombre

    actividad = ChecklistEstadoController.listar_actividad(licitacion_id, db, limit)
    for entrada in actividad:
        entrada["item_nombre"] = nombres_por_key.get(entrada["item_key"], entrada["item_key"])

    return actividad


def build_licitacion_explorer(licitacion_id: str, db: Session) -> Dict[str, Any]:
    licitacion = db.query(Licitacion).filter(Licitacion.id == licitacion_id).first()
    if not licitacion:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Licitacion no encontrada")

    empresa = db.query(Empresa).filter(Empresa.id == licitacion.empresa_id).first()
    if not empresa:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Empresa no encontrada")

    # Solo las carpetas/documentos de ESTA licitacion, no de toda la empresa.
    folders = db.query(Carpeta).filter(Carpeta.licitacion_id == licitacion.id).order_by(Carpeta.nombre.asc()).all()
    folder_ids = [folder.id for folder in folders]
    documents = (
        db.query(Documento)
        .filter(Documento.carpeta_id.in_(folder_ids))
        .order_by(Documento.created_at.desc())
        .all()
        if folder_ids
        else []
    )

    from src.controllers.licitacion_controller import LicitacionController

    LicitacionController._avanzar_fase_por_cronograma(licitacion, db)

    tree = build_folder_tree(folders, documents)
    required_documents = build_checklist_items(licitacion.id, documents, db, licitacion.checklist_excluidos)
    obligatory_completed = sum(1 for item in required_documents if item["cumple"] and item["obligatorio"])
    obligatory_total = sum(1 for item in required_documents if item["obligatorio"])
    semaforo = LicitacionController.calcular_semaforo(licitacion, db)

    pliego_analysis = None
    if licitacion.pliego_texto:
        pliego_analysis = analyze_text(licitacion.pliego_texto)

    rup_analysis = None
    if licitacion.rup_texto:
        rup_analysis = analyze_text(licitacion.rup_texto)

    comparativo_codigos = comparar_codigos_pliego_rup(licitacion.pliego_texto, licitacion.rup_texto, licitacion.objeto_contrato)
    indicadores_financieros = comparar_indicadores_financieros(
        licitacion.rup_texto,
        licitacion.pliego_texto,
        licitacion.indicadores_financieros_requeridos,
        licitacion.indicadores_financieros_rup_manual,
    )

    cuantia = float(licitacion.cuantia) if licitacion.cuantia is not None else 0.0
    calculos = {
        "valor_base": cuantia,
        "iva_19": round(cuantia * 0.19, 2),
        "aiu_10": round(cuantia * 0.10, 2),
        "subtotal_con_iva": round(cuantia * 1.19, 2),
        "total_con_aiu": round(cuantia * 1.10, 2),
        "retencion_2": round(cuantia * 0.02, 2),
    }

    return {
        "licitacion": {**serialize_licitacion(licitacion), "semaforo": semaforo},
        "empresa": serialize_empresa(empresa),
        "carpetas": tree["arbol"],
        "documentos_raiz": tree["documentos_raiz"],
        "documentos": [serialize_documento(document) for document in documents],
        "documentos_obligatorios": required_documents,
        "resumen_documental": {
            "total_documentos": len(documents),
            "documentos_vigentes": sum(1 for document in documents if document.vigente),
            "total_carpetas": len(folders),
            "obligatorios_total": obligatory_total,
            "obligatorios_cumplidos": obligatory_completed,
            "obligatorios_pendientes": max(obligatory_total - obligatory_completed, 0),
            "cobertura_porcentaje": round((obligatory_completed / obligatory_total) * 100, 1) if obligatory_total else 0,
        },
        "pliego_analisis": pliego_analysis,
        "rup_analisis": rup_analysis,
        "comparativo_codigos": comparativo_codigos,
        "indicadores_financieros": indicadores_financieros,
        "calculos": calculos,
        "secop": {
            "numero_secop": licitacion.numero_secop,
            "url_secop": licitacion.url_secop,
            "entidad_contratante": licitacion.entidad_contratante,
            "nit_entidad": licitacion.nit_entidad,
            "objeto_contrato": licitacion.objeto_contrato,
            "cuantia": cuantia,
            "estado": licitacion.estado,
            "fecha_cierre": licitacion.fecha_cierre.isoformat() if licitacion.fecha_cierre else None,
            "dias_restantes": serialize_licitacion(licitacion)["dias_restantes"],
        },
    }
