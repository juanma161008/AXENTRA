import io
import re
import unicodedata
import uuid
from typing import Optional

from fastapi import HTTPException, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from src.models.entidad import Entidad
from src.schemas.entidad import EntidadCreate, EntidadUpdate
from src.utils.constants import MENSAJES


NOMBRE_ALIASES = {"nombre", "entidad", "razon", "empresa", "contratante"}
NIT_ALIASES = {"nit", "identificacion", "documento", "cedula", "rut", "cc"}
FILAS_MAX_BUSQUEDA_ENCABEZADO = 20


def _normalizar_encabezado(valor) -> str:
    texto = str(valor or "").strip().lower()
    sin_acentos = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return sin_acentos


def _palabras(texto: str) -> set:
    return set(re.split(r"[^a-z0-9]+", texto)) - {""}


def _normalizar_nit(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    texto = str(valor).strip()
    return re.sub(r"[^0-9A-Za-z-]", "", texto)


class EntidadController:
    @staticmethod
    def get_entidades(db: Session, skip: int = 0, limit: int = 100, activo: Optional[bool] = None, q: Optional[str] = None):
        query = db.query(Entidad)
        if activo is not None:
            query = query.filter(Entidad.activo == activo)
        if q:
            like = f"%{q}%"
            query = query.filter((Entidad.nombre.ilike(like)) | (Entidad.nit.ilike(like)))
        return query.order_by(Entidad.nombre.asc()).offset(skip).limit(limit).all()

    @staticmethod
    def get_entidad(entidad_id: uuid.UUID, db: Session) -> Entidad:
        entidad = db.query(Entidad).filter(Entidad.id == entidad_id).first()
        if not entidad:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Entidad no encontrada")
        return entidad

    @staticmethod
    def buscar_por_nit(nit: str, db: Session) -> Optional[Entidad]:
        if not nit:
            return None
        return db.query(Entidad).filter(Entidad.nit == nit).first()

    @staticmethod
    def get_or_create_by_nit(nit: Optional[str], nombre: Optional[str], db: Session) -> Optional[Entidad]:
        """Reutiliza la entidad si el NIT ya existe; si no, la crea a partir de los datos de la licitacion."""
        if not nit or not nombre:
            return None

        existente = EntidadController.buscar_por_nit(nit, db)
        if existente:
            return existente

        nueva = Entidad(id=uuid.uuid4(), nombre=nombre, nit=nit)
        db.add(nueva)
        db.commit()
        db.refresh(nueva)
        return nueva

    @staticmethod
    def create_entidad(data: EntidadCreate, db: Session) -> Entidad:
        existing = db.query(Entidad).filter(Entidad.nit == data.nit).first()
        if existing:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El NIT ya esta registrado")

        nueva = Entidad(id=uuid.uuid4(), **data.model_dump(), activo=True)
        db.add(nueva)
        db.commit()
        db.refresh(nueva)
        return nueva

    @staticmethod
    def update_entidad(entidad_id: uuid.UUID, data: EntidadUpdate, db: Session) -> Entidad:
        entidad = EntidadController.get_entidad(entidad_id, db)

        if data.nit and data.nit != entidad.nit:
            existing = db.query(Entidad).filter(Entidad.nit == data.nit).first()
            if existing:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El NIT ya esta registrado por otra entidad")

        update_data = data.model_dump(exclude_unset=True)
        for key, value in update_data.items():
            setattr(entidad, key, value)

        db.commit()
        db.refresh(entidad)
        return entidad

    @staticmethod
    def import_from_excel(file_bytes: bytes, db: Session) -> dict:
        """Importa entidades desde un Excel buscando las columnas 'nombre' y 'nit'."""
        try:
            workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        except Exception:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No fue posible leer el archivo. Verifica que sea un Excel válido (.xlsx)",
            )

        hoja = workbook.active
        filas = hoja.iter_rows(values_only=True)

        columna_nombre = None
        columna_nit = None

        for indice_fila, fila in enumerate(filas):
            if fila is None or all(valor is None for valor in fila):
                continue

            normalizados = [_normalizar_encabezado(valor) for valor in fila]
            candidato_nombre = next(
                (i for i, texto in enumerate(normalizados) if NOMBRE_ALIASES & _palabras(texto)), None
            )
            candidato_nit = next(
                (i for i, texto in enumerate(normalizados) if NIT_ALIASES & _palabras(texto)), None
            )

            if candidato_nombre is not None and candidato_nit is not None and candidato_nombre != candidato_nit:
                columna_nombre, columna_nit = candidato_nombre, candidato_nit
                break

            if indice_fila + 1 >= FILAS_MAX_BUSQUEDA_ENCABEZADO:
                break

        if columna_nombre is None or columna_nit is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No se encontraron columnas de Nombre y NIT en el Excel. Deben estar entre las primeras 20 filas de la hoja activa.",
            )

        creadas = 0
        actualizadas = 0
        omitidas = 0
        errores = []
        nits_en_archivo = set()

        for numero_fila, fila in enumerate(filas, start=indice_fila + 2):
            if fila is None or all(valor is None for valor in fila):
                continue

            nombre = str(fila[columna_nombre]).strip() if columna_nombre < len(fila) and fila[columna_nombre] is not None else ""
            nit = _normalizar_nit(fila[columna_nit] if columna_nit < len(fila) else None)

            if not nombre or not nit:
                omitidas += 1
                errores.append(f"Fila {numero_fila}: falta nombre o NIT")
                continue

            if nit in nits_en_archivo:
                omitidas += 1
                errores.append(f"Fila {numero_fila}: NIT {nit} repetido en el archivo")
                continue
            nits_en_archivo.add(nit)

            existente = db.query(Entidad).filter(Entidad.nit == nit).first()
            if existente:
                if existente.nombre != nombre:
                    existente.nombre = nombre
                    actualizadas += 1
            else:
                db.add(Entidad(id=uuid.uuid4(), nombre=nombre, nit=nit, activo=True))
                creadas += 1

        db.commit()

        return {
            "creadas": creadas,
            "actualizadas": actualizadas,
            "omitidas": omitidas,
            "errores": errores[:20],
        }

    @staticmethod
    def delete_entidad(entidad_id: uuid.UUID, db: Session):
        entidad = EntidadController.get_entidad(entidad_id, db)
        entidad.activo = False
        db.commit()
        return {"message": MENSAJES.get("ENTIDAD_ELIMINADA", "Entidad desactivada correctamente")}
