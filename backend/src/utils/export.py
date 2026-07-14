import csv
import io
from typing import List, Dict, Any, Optional
from datetime import datetime

def export_to_csv(data: List[Dict[str, Any]], filename: str = "export.csv") -> bytes:
    """Exportar datos a CSV"""
    output = io.StringIO()
    if not data:
        return "".encode('utf-8')
    
    writer = csv.DictWriter(output, fieldnames=data[0].keys())
    writer.writeheader()
    writer.writerows(data)
    return output.getvalue().encode('utf-8-sig')

def export_to_excel(data: List[Dict[str, Any]]) -> bytes:
    """Exportar datos a Excel (XLSX)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Datos"
        
        if not data:
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()
        
        # Headers
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for row, item in enumerate(data, 2):
            for col, key in enumerate(headers, 1):
                value = item.get(key, "")
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M")
                ws.cell(row=row, column=col, value=value)
        
        # Autoajustar columnas
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col[0].column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
        
    except ImportError:
        # Fallback a CSV si openpyxl no está instalado
        return export_to_csv(data)